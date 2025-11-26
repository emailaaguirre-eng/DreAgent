"""
Example Custom Tasks for Lea
Copy and modify these to create your own tasks
"""

from lea_tasks import BaseTask, TaskResult
import logging
import os
import json
from pathlib import Path

# =====================================================
# EXAMPLE TASK 1: Text Analysis Task
# =====================================================

class TextAnalysisTask(BaseTask):
    """Analyze text file and extract key information"""
    
    def __init__(self):
        super().__init__(
            name="text_analyze",
            description="Analyze a text file for word count, reading time, etc.",
            requires_confirmation=False
        )
    
    def get_required_params(self):
        return ["file_path"]
    
    def validate_params(self, **kwargs):
        file_path = kwargs.get("file_path")
        if not file_path:
            return False, "File path is required"
        if not os.path.exists(file_path):
            return False, f"File does not exist: {file_path}"
        return True, ""
    
    def execute(self, **kwargs):
        try:
            file_path = kwargs.get("file_path")
            
            # Validate
            valid, msg = self.validate_params(**kwargs)
            if not valid:
                return TaskResult(False, msg, error=msg)
            
            # Read file
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Analyze
            words = content.split()
            word_count = len(words)
            char_count = len(content)
            reading_time_minutes = word_count / 200  # Average reading speed
            
            # Find most common words
            from collections import Counter
            word_counts = Counter(word.lower() for word in words if len(word) > 3)
            top_words = dict(word_counts.most_common(10))
            
            analysis = {
                "file_path": file_path,
                "word_count": word_count,
                "character_count": char_count,
                "reading_time_minutes": round(reading_time_minutes, 2),
                "top_words": top_words
            }
            
            return TaskResult(
                True,
                f"Analyzed {file_path}: {word_count} words, ~{reading_time_minutes:.1f} min read",
                data=analysis
            )
        except Exception as e:
            logging.error(f"TextAnalysisTask error: {e}")
            return TaskResult(False, f"Failed to analyze text: {str(e)}", error=str(e))


# =====================================================
# EXAMPLE TASK 2: Configuration Manager
# =====================================================

class ConfigManagerTask(BaseTask):
    """Read or update configuration files"""
    
    def __init__(self):
        super().__init__(
            name="config_manager",
            description="Read or update JSON configuration files",
            requires_confirmation=True  # Config changes should be confirmed
        )
    
    def get_required_params(self):
        return ["config_path", "action"]
    
    def validate_params(self, **kwargs):
        action = kwargs.get("action", "").lower()
        if action not in ["read", "set", "get"]:
            return False, "Action must be 'read', 'set', or 'get'"
        return True, ""
    
    def execute(self, **kwargs):
        try:
            config_path = kwargs.get("config_path")
            action = kwargs.get("action", "read").lower()
            key = kwargs.get("key")  # For get/set operations
            value = kwargs.get("value")  # For set operations
            
            if not os.path.exists(config_path):
                # Create empty config if it doesn't exist
                if action == "read":
                    return TaskResult(False, f"Config file does not exist: {config_path}")
                with open(config_path, 'w') as f:
                    json.dump({}, f, indent=2)
            
            # Read config
            with open(config_path, 'r') as f:
                config = json.load(f)
            
            if action == "read":
                return TaskResult(
                    True,
                    f"Read config from {config_path}",
                    data={"config": config}
                )
            
            elif action == "get":
                if not key:
                    return TaskResult(False, "Key required for 'get' action")
                result = config.get(key)
                return TaskResult(
                    True,
                    f"Got {key} = {result}",
                    data={"key": key, "value": result}
                )
            
            elif action == "set":
                if not key:
                    return TaskResult(False, "Key and value required for 'set' action")
                
                # Update config
                config[key] = value
                
                # Write back
                with open(config_path, 'w') as f:
                    json.dump(config, f, indent=2)
                
                return TaskResult(
                    True,
                    f"Set {key} = {value} in {config_path}",
                    data={"key": key, "value": value, "config": config}
                )
            
        except Exception as e:
            logging.error(f"ConfigManagerTask error: {e}")
            return TaskResult(False, f"Config operation failed: {str(e)}", error=str(e))


# =====================================================
# EXAMPLE TASK 3: File Organizer
# =====================================================

class FileOrganizerTask(BaseTask):
    """Organize files by extension or date"""
    
    def __init__(self):
        super().__init__(
            name="file_organize",
            description="Organize files in a directory by extension or date",
            requires_confirmation=True  # Moving files requires confirmation
        )
    
    def get_required_params(self):
        return ["directory", "organize_by"]
    
    def validate_params(self, **kwargs):
        directory = kwargs.get("directory")
        organize_by = kwargs.get("organize_by", "").lower()
        
        if not directory or not os.path.exists(directory):
            return False, f"Directory does not exist: {directory}"
        if organize_by not in ["extension", "date"]:
            return False, "organize_by must be 'extension' or 'date'"
        return True, ""
    
    def execute(self, **kwargs):
        try:
            from lea_tasks import get_task_registry
            from datetime import datetime
            
            directory = kwargs.get("directory")
            organize_by = kwargs.get("organize_by", "extension").lower()
            
            registry = get_task_registry()
            
            # List directory
            list_result = registry.execute_task(
                "directory_list",
                {"path": directory},
                confirmed=False
            )
            
            if not list_result.success:
                return TaskResult(False, "Failed to list directory", error=list_result.error)
            
            files = [item for item in list_result.data['items'] if item['type'] == 'file']
            moved_count = 0
            
            for file_item in files:
                file_path = file_item['path']
                file_name = file_item['name']
                
                if organize_by == "extension":
                    # Get extension
                    ext = file_name.split('.')[-1] if '.' in file_name else "no_extension"
                    target_dir = os.path.join(directory, ext)
                    
                elif organize_by == "date":
                    # Get file modification date
                    mod_time = os.path.getmtime(file_path)
                    mod_date = datetime.fromtimestamp(mod_time)
                    date_folder = mod_date.strftime("%Y-%m-%d")
                    target_dir = os.path.join(directory, date_folder)
                
                # Create target directory
                registry.execute_task(
                    "directory_create",
                    {"path": target_dir},
                    confirmed=False
                )
                
                # Move file
                target_path = os.path.join(target_dir, file_name)
                move_result = registry.execute_task(
                    "file_move",
                    {"source": file_path, "destination": target_path},
                    confirmed=False  # Already confirmed for the whole operation
                )
                
                if move_result.success:
                    moved_count += 1
            
            return TaskResult(
                True,
                f"Organized {moved_count} files by {organize_by}",
                data={"moved_count": moved_count, "organize_by": organize_by}
            )
            
        except Exception as e:
            logging.error(f"FileOrganizerTask error: {e}")
            return TaskResult(False, f"Organization failed: {str(e)}", error=str(e))


# =====================================================
# EMAIL TASK: Outlook Email Checker (Executive Assistant Only)
# =====================================================

def _sanitize_credentials(text, client_id=None, client_secret=None, tenant_id=None):
    """Remove credential information from text to prevent exposure in logs"""
    if not text:
        return text
    sanitized = str(text)
    if client_id:
        sanitized = sanitized.replace(client_id, "[CLIENT_ID_REDACTED]")
    if client_secret:
        sanitized = sanitized.replace(client_secret, "[CLIENT_SECRET_REDACTED]")
    if tenant_id:
        sanitized = sanitized.replace(tenant_id, "[TENANT_ID_REDACTED]")
    return sanitized

class OutlookEmailCheckTask(BaseTask):
    """Check Outlook inbox emails and generate Excel report (Executive Assistant mode only)"""
    
    def __init__(self):
        super().__init__(
            name="outlook_email_check",
            description="Check Outlook inbox emails and generate Excel report (Executive Assistant mode only)",
            requires_confirmation=False
        )
    
    def get_required_params(self):
        return []  # No required params - uses .env OAuth credentials
    
    def execute(self, **kwargs):
        # Initialize credential variables outside try block for sanitization
        client_id = None
        client_secret = None
        tenant_id = None
        
        # Status messages (no email content exposed)
        status_messages = []
        
        try:
            import os
            from dotenv import load_dotenv
            load_dotenv()
            
            status_messages.append("🔐 Step 1/5: Loading authentication libraries...")
            
            # Try to import Microsoft Graph SDK
            try:
                from msal import ConfidentialClientApplication
                import requests
                import logging as msal_logging
                # Suppress MSAL verbose logging to prevent credential exposure
                msal_logging.getLogger("msal").setLevel(msal_logging.CRITICAL)  # Only critical errors
                # Also suppress requests logging
                logging.getLogger("requests").setLevel(logging.WARNING)
                logging.getLogger("urllib3").setLevel(logging.WARNING)
            except ImportError:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Microsoft authentication libraries not installed.\nInstall with: pip install msal requests",
                    error="Missing dependencies"
                )
            
            status_messages.append("✅ Step 1/5: Libraries loaded")
            status_messages.append("🔐 Step 2/5: Reading OAuth credentials...")
            
            # Get OAuth credentials from .env
            client_id = os.getenv("OUTLOOK_CLIENT_ID")
            client_secret = os.getenv("OUTLOOK_CLIENT_SECRET")
            tenant_id = os.getenv("OUTLOOK_TENANT_ID", "common")  # Default to 'common' for personal accounts
            
            if not client_id or not client_secret:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Outlook OAuth credentials not found in .env file.\nNeed: OUTLOOK_CLIENT_ID, OUTLOOK_CLIENT_SECRET, and optionally OUTLOOK_TENANT_ID",
                    error="Missing credentials"
                )
            
            status_messages.append("✅ Step 2/5: Credentials found")
            status_messages.append("🔐 Step 3/5: Authenticating with Outlook...")
            
            # Authenticate
            authority = f"https://login.microsoftonline.com/{tenant_id}"
            app = ConfidentialClientApplication(
                client_id=client_id,
                client_credential=client_secret,
                authority=authority
            )
            
            # Get token (using client credentials flow - may need user consent first)
            scopes = ["https://graph.microsoft.com/.default"]
            result = app.acquire_token_silent(scopes, account=None)
            
            if not result:
                # Need user interaction - return message to user
                auth_url = app.get_authorization_request_url(
                    scopes=["https://graph.microsoft.com/Mail.Read"],
                    redirect_uri="http://localhost:8000"  # Placeholder
                )
                return TaskResult(
                    False,
                    "⚠️ Status: Authentication required.\nPlease authenticate Outlook access when prompted.",
                    error="Authentication required",
                    data={"auth_url": auth_url, "needs_auth": True}
                )
            
            access_token = result.get("access_token")
            if not access_token:
                error_desc = result.get('error_description', 'Unknown error')
                # Sanitize error description
                error_desc = _sanitize_credentials(error_desc, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Authentication error.\n{error_desc}",
                    error="Authentication failed"
                )
            
            status_messages.append("✅ Step 3/5: Authenticated successfully")
            status_messages.append("📧 Step 4/5: Connecting to inbox...")
            
            # Get inbox messages (not folders/subfolders - just the inbox)
            headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json"
            }
            
            # Get inbox folder ID first
            inbox_url = "https://graph.microsoft.com/v1.0/me/mailFolders/inbox"
            response = requests.get(inbox_url, headers=headers)
            
            if response.status_code != 200:
                # Sanitize response text to prevent credential exposure
                response_text = _sanitize_credentials(response.text, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Could not access inbox (Error {response.status_code}).\nPlease check your permissions.",
                    error=f"API error: {response.status_code}"
                )
            
            inbox_data = response.json()
            inbox_id = inbox_data.get("id")
            
            status_messages.append("📥 Step 5/5: Retrieving email data...")
            
            # Get messages from inbox only
            messages_url = f"https://graph.microsoft.com/v1.0/me/mailFolders/{inbox_id}/messages"
            params = {
                "$top": 50,  # Get up to 50 most recent
                "$orderby": "receivedDateTime desc",
                "$select": "subject,sender,receivedDateTime,isRead,hasAttachments,bodyPreview"
            }
            
            response = requests.get(messages_url, headers=headers, params=params)
            
            if response.status_code != 200:
                # Sanitize response text to prevent credential exposure
                response_text = _sanitize_credentials(response.text, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Could not retrieve emails (Error {response.status_code}).\nPlease check your permissions.",
                    error=f"API error: {response.status_code}"
                )
            
            messages_data = response.json()
            messages = messages_data.get("value", [])
            
            # Count unread (don't log individual email info)
            unread_count = sum(1 for msg in messages if not msg.get("isRead", False))
            total_count = len(messages)
            
            status_messages.append(f"✅ Step 5/5: Retrieved {total_count} emails ({unread_count} unread)")
            status_messages.append("📊 Generating Excel report...")
            
            # Try to create Excel file
            try:
                try:
                    import openpyxl
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                except ImportError:
                    # Fallback to pandas if openpyxl not available
                    try:
                        import pandas as pd
                        use_pandas = True
                    except ImportError:
                        return TaskResult(
                            False,
                            "❌ Status: Excel library not installed.\nInstall with: pip install openpyxl OR pip install pandas openpyxl",
                            error="Missing dependencies"
                        )
                else:
                    use_pandas = False
                
                # Get reports directory (F:\Dre_Programs\LeaAssistant\lea_reports)
                from pathlib import Path
                project_dir = Path("F:/Dre_Programs/LeaAssistant")
                reports_dir = project_dir / "lea_reports"
                reports_dir.mkdir(parents=True, exist_ok=True)
                
                # Create filename with timestamp
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"outlook_inbox_report_{timestamp}.xlsx"
                file_path = reports_dir / filename
                
                if use_pandas:
                    # Use pandas to create Excel
                    email_data = []
                    for msg in messages:
                        sender = msg.get("sender", {}).get("emailAddress", {}).get("address", "Unknown")
                        sender_name = msg.get("sender", {}).get("emailAddress", {}).get("name", "")
                        subject = msg.get("subject", "(No Subject)")
                        received = msg.get("receivedDateTime", "")
                        is_read = msg.get("isRead", False)
                        has_attachments = msg.get("hasAttachments", False)
                        preview = msg.get("bodyPreview", "")[:500]  # First 500 chars
                        
                        # Format date
                        try:
                            dt = datetime.fromisoformat(received.replace('Z', '+00:00'))
                            date_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                        except:
                            date_str = received
                        
                        email_data.append({
                            "Status": "Read" if is_read else "Unread",
                            "Subject": subject,
                            "From Name": sender_name,
                            "From Email": sender,
                            "Received": date_str,
                            "Has Attachments": "Yes" if has_attachments else "No",
                            "Preview": preview
                        })
                    
                    df = pd.DataFrame(email_data)
                    df.to_excel(file_path, index=False, sheet_name="Inbox Emails")
                    
                else:
                    # Use openpyxl for more control
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Inbox Emails"
                    
                    # Header row with styling
                    headers = ["Status", "Subject", "From Name", "From Email", "Received", "Has Attachments", "Preview"]
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Add summary row
                    ws.cell(row=2, column=1, value="Summary")
                    ws.cell(row=2, column=2, value=f"Total: {total_count} | Unread: {unread_count} | Read: {total_count - unread_count}")
                    ws.merge_cells(f"B2:G2")
                    
                    # Add email data
                    for row_num, msg in enumerate(messages, start=3):
                        sender = msg.get("sender", {}).get("emailAddress", {}).get("address", "Unknown")
                        sender_name = msg.get("sender", {}).get("emailAddress", {}).get("name", "")
                        subject = msg.get("subject", "(No Subject)")
                        received = msg.get("receivedDateTime", "")
                        is_read = msg.get("isRead", False)
                        has_attachments = msg.get("hasAttachments", False)
                        preview = msg.get("bodyPreview", "")[:500]  # First 500 chars
                        
                        # Format date
                        try:
                            dt = datetime.fromisoformat(received.replace('Z', '+00:00'))
                            date_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                        except:
                            date_str = received
                        
                        ws.cell(row=row_num, column=1, value="Read" if is_read else "Unread")
                        ws.cell(row=row_num, column=2, value=subject)
                        ws.cell(row=row_num, column=3, value=sender_name)
                        ws.cell(row=row_num, column=4, value=sender)
                        ws.cell(row=row_num, column=5, value=date_str)
                        ws.cell(row=row_num, column=6, value="Yes" if has_attachments else "No")
                        ws.cell(row=row_num, column=7, value=preview)
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        ws.column_dimensions[col_letter].width = adjusted_width
                    
                    wb.save(file_path)
                
                status_messages.append("✅ Excel report generated successfully")
                
                # Generate summary text for message (NO EMAIL CONTENT - only status and file location)
                summary_text = "\n".join(status_messages)
                summary_text += f"\n\n✅ Report Complete!\n"
                summary_text += f"📊 Summary: {total_count} total emails ({unread_count} unread, {total_count - unread_count} read)\n"
                summary_text += f"📁 Excel file saved to:\n   {file_path}\n"
                summary_text += f"📝 Report contains {len(messages)} email entries.\n"
                summary_text += f"\n💡 All email details are in the Excel file - no email content is shown in chat for security."
                
                return TaskResult(
                    True,
                    summary_text,  # Only status, no email content
                    data={
                        "total_count": total_count,
                        "unread_count": unread_count,
                        "read_count": total_count - unread_count,
                        "messages": messages,  # Keep in data but not in message/logs
                        "file_path": str(file_path),
                        "filename": filename
                    }
                )
                
            except Exception as excel_error:
                # If Excel creation fails, return error (don't expose email content in fallback)
                error_msg = _sanitize_credentials(str(excel_error), client_id, client_secret, tenant_id)
                # Don't log email content - just log that Excel failed
                
                status_messages.append("❌ Excel generation failed")
                
                return TaskResult(
                    False,
                    "\n".join(status_messages) + f"\n\n❌ Failed to generate Excel file.\nError: {error_msg}\n\nInstall openpyxl with: pip install openpyxl",
                    data={
                        "total_count": total_count,
                        "unread_count": unread_count,
                        "read_count": total_count - unread_count,
                        "excel_error": error_msg
                    },
                    error=error_msg
                )
            
        except Exception as e:
            # Sanitize error message to remove any credential information
            error_msg = _sanitize_credentials(str(e), client_id, client_secret, tenant_id)
            
            # Log error without credentials or email content (just status)
            # Don't log the actual error message as it might contain email data
            logging.error("OutlookEmailCheckTask error occurred (details not logged for security)")
            
            # Return sanitized error with status (don't include full traceback as it might contain credentials/email data)
            status_text = "\n".join(status_messages) if status_messages else "❌ Status: Task failed"
            return TaskResult(
                False,
                f"{status_text}\n\n❌ Error: {error_msg}\n\nPlease check your OAuth credentials and try again.",
                error=error_msg  # Don't include traceback to avoid credential/email exposure
            )

class OutlookEmailDraftTask(BaseTask):
    """Create a draft email in Outlook (Executive Assistant mode only)"""
    
    def __init__(self):
        super().__init__(
            name="outlook_email_draft",
            description="Create a draft email in Outlook with subject, body, and optional recipients (Executive Assistant mode only)",
            requires_confirmation=False
        )
    
    def get_required_params(self):
        return ["subject", "body"]
    
    def execute(self, **kwargs):
        # Initialize credential variables outside try block for sanitization
        client_id = None
        client_secret = None
        tenant_id = None
        
        # Status messages (no email content exposed)
        status_messages = []
        
        try:
            import os
            from dotenv import load_dotenv
            load_dotenv()
            
            status_messages.append("🔐 Step 1/4: Loading authentication libraries...")
            
            # Try to import Microsoft Graph SDK
            try:
                from msal import ConfidentialClientApplication
                import requests
                import logging as msal_logging
                # Suppress MSAL verbose logging to prevent credential exposure
                msal_logging.getLogger("msal").setLevel(msal_logging.CRITICAL)
                logging.getLogger("requests").setLevel(logging.WARNING)
                logging.getLogger("urllib3").setLevel(logging.WARNING)
            except ImportError:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Microsoft authentication libraries not installed.\nInstall with: pip install msal requests",
                    error="Missing dependencies"
                )
            
            status_messages.append("✅ Step 1/4: Libraries loaded")
            status_messages.append("🔐 Step 2/4: Reading OAuth credentials...")
            
            # Get OAuth credentials from .env
            client_id = os.getenv("OUTLOOK_CLIENT_ID")
            client_secret = os.getenv("OUTLOOK_CLIENT_SECRET")
            tenant_id = os.getenv("OUTLOOK_TENANT_ID", "common")
            
            if not client_id or not client_secret:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Outlook OAuth credentials not found in .env file.\nNeed: OUTLOOK_CLIENT_ID, OUTLOOK_CLIENT_SECRET, and optionally OUTLOOK_TENANT_ID",
                    error="Missing credentials"
                )
            
            status_messages.append("✅ Step 2/4: Credentials found")
            status_messages.append("🔐 Step 3/4: Authenticating with Outlook...")
            
            # Authenticate
            authority = f"https://login.microsoftonline.com/{tenant_id}"
            app = ConfidentialClientApplication(
                client_id=client_id,
                client_credential=client_secret,
                authority=authority
            )
            
            # Get token
            scopes = ["https://graph.microsoft.com/.default"]
            result = app.acquire_token_silent(scopes, account=None)
            
            if not result:
                return TaskResult(
                    False,
                    "⚠️ Status: Authentication required.\nPlease authenticate Outlook access when prompted.",
                    error="Authentication required",
                    data={"needs_auth": True}
                )
            
            access_token = result.get("access_token")
            if not access_token:
                error_desc = result.get('error_description', 'Unknown error')
                error_desc = _sanitize_credentials(error_desc, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Authentication error.\n{error_desc}",
                    error="Authentication failed"
                )
            
            status_messages.append("✅ Step 3/4: Authenticated successfully")
            status_messages.append("📧 Step 4/4: Creating draft email...")
            
            # Get email parameters
            subject = kwargs.get("subject", "")
            body = kwargs.get("body", "")
            to_recipients = kwargs.get("to", [])  # List of email addresses
            cc_recipients = kwargs.get("cc", [])  # List of email addresses
            bcc_recipients = kwargs.get("bcc", [])  # List of email addresses
            
            if not subject or not body:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Subject and body are required for draft email.",
                    error="Missing required parameters"
                )
            
            # Build recipients
            message = {
                "subject": subject,
                "body": {
                    "contentType": "HTML",
                    "content": body.replace("\n", "<br>")  # Convert newlines to HTML breaks
                }
            }
            
            # Add recipients if provided
            if to_recipients:
                if isinstance(to_recipients, str):
                    to_recipients = [to_recipients]
                message["toRecipients"] = [{"emailAddress": {"address": addr}} for addr in to_recipients]
            
            if cc_recipients:
                if isinstance(cc_recipients, str):
                    cc_recipients = [cc_recipients]
                message["ccRecipients"] = [{"emailAddress": {"address": addr}} for addr in cc_recipients]
            
            if bcc_recipients:
                if isinstance(bcc_recipients, str):
                    bcc_recipients = [bcc_recipients]
                message["bccRecipients"] = [{"emailAddress": {"address": addr}} for addr in bcc_recipients]
            
            # Create draft via Microsoft Graph API
            headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json"
            }
            
            draft_url = "https://graph.microsoft.com/v1.0/me/messages"
            response = requests.post(draft_url, headers=headers, json=message)
            
            if response.status_code not in [201, 200]:
                response_text = _sanitize_credentials(response.text, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Could not create draft (Error {response.status_code}).\nPlease check your permissions.",
                    error=f"API error: {response.status_code}"
                )
            
            draft_data = response.json()
            draft_id = draft_data.get("id")
            
            status_messages.append("✅ Step 4/4: Draft email created successfully")
            
            # Generate summary (no sensitive content)
            summary_text = "\n".join(status_messages)
            summary_text += f"\n\n✅ Draft Created!\n"
            summary_text += f"📧 Subject: {subject}\n"
            if to_recipients:
                summary_text += f"📬 To: {', '.join(to_recipients) if isinstance(to_recipients, list) else to_recipients}\n"
            if cc_recipients:
                summary_text += f"📋 CC: {', '.join(cc_recipients) if isinstance(cc_recipients, list) else cc_recipients}\n"
            if bcc_recipients:
                summary_text += f"🔒 BCC: {len(bcc_recipients) if isinstance(bcc_recipients, list) else 1} recipient(s)\n"
            summary_text += f"📝 Draft ID: {draft_id}\n"
            summary_text += f"\n💡 Draft is ready in your Outlook drafts folder. You can review and send it when ready."
            
            return TaskResult(
                True,
                summary_text,
                data={
                    "draft_id": draft_id,
                    "subject": subject,
                    "to_recipients": to_recipients,
                    "cc_recipients": cc_recipients,
                    "bcc_recipients": bcc_recipients
                }
            )
            
        except Exception as e:
            # Sanitize error message
            error_msg = _sanitize_credentials(str(e), client_id, client_secret, tenant_id)
            logging.error("OutlookEmailDraftTask error occurred (details not logged for security)")
            
            status_text = "\n".join(status_messages) if status_messages else "❌ Status: Task failed"
            return TaskResult(
                False,
                f"{status_text}\n\n❌ Error: {error_msg}\n\nPlease check your OAuth credentials and try again.",
                error=error_msg
            )


# =====================================================
# REGISTER CUSTOM TASKS
# =====================================================

class OutlookCalendarCheckTask(BaseTask):
    """Check Outlook calendar events and generate Excel report (Executive Assistant mode only) - READ ONLY"""
    
    def __init__(self):
        super().__init__(
            name="outlook_calendar_check",
            description="Check Outlook calendar events and generate a report (Executive Assistant mode only) - READ ONLY",
            requires_confirmation=False
        )
    
    def get_required_params(self):
        return []  # No required params - uses .env OAuth credentials
    
    def execute(self, **kwargs):
        # Initialize credential variables outside try block for sanitization
        client_id = None
        client_secret = None
        tenant_id = None
        
        # Status messages
        status_messages = []
        
        try:
            import os
            from dotenv import load_dotenv
            load_dotenv()
            
            status_messages.append("🔐 Step 1/5: Loading authentication libraries...")
            
            # Try to import Microsoft Graph SDK
            try:
                from msal import ConfidentialClientApplication
                import requests
                import logging as msal_logging
                # Suppress MSAL verbose logging to prevent credential exposure
                msal_logging.getLogger("msal").setLevel(msal_logging.CRITICAL)
                logging.getLogger("requests").setLevel(logging.WARNING)
                logging.getLogger("urllib3").setLevel(logging.WARNING)
            except ImportError:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Microsoft authentication libraries not installed.\nInstall with: pip install msal requests",
                    error="Missing dependencies"
                )
            
            status_messages.append("✅ Step 1/5: Libraries loaded")
            status_messages.append("🔐 Step 2/5: Reading OAuth credentials...")
            
            # Get OAuth credentials from .env
            client_id = os.getenv("OUTLOOK_CLIENT_ID")
            client_secret = os.getenv("OUTLOOK_CLIENT_SECRET")
            tenant_id = os.getenv("OUTLOOK_TENANT_ID", "common")
            
            if not client_id or not client_secret:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Outlook OAuth credentials not found in .env file.\nNeed: OUTLOOK_CLIENT_ID, OUTLOOK_CLIENT_SECRET, and optionally OUTLOOK_TENANT_ID",
                    error="Missing credentials"
                )
            
            status_messages.append("✅ Step 2/5: Credentials found")
            status_messages.append("🔐 Step 3/5: Authenticating with Outlook...")
            
            # Authenticate
            authority = f"https://login.microsoftonline.com/{tenant_id}"
            app = ConfidentialClientApplication(
                client_id=client_id,
                client_credential=client_secret,
                authority=authority
            )
            
            # Get token
            scopes = ["https://graph.microsoft.com/.default"]
            result = app.acquire_token_silent(scopes, account=None)
            
            if not result:
                return TaskResult(
                    False,
                    "⚠️ Status: Authentication required.\nPlease authenticate Outlook access when prompted.",
                    error="Authentication required",
                    data={"needs_auth": True}
                )
            
            access_token = result.get("access_token")
            if not access_token:
                error_desc = result.get('error_description', 'Unknown error')
                error_desc = _sanitize_credentials(error_desc, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Authentication error.\n{error_desc}",
                    error="Authentication failed"
                )
            
            status_messages.append("✅ Step 3/5: Authenticated successfully")
            status_messages.append("📅 Step 4/5: Retrieving calendar events...")
            
            # Get calendar events (next 30 days by default)
            headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json"
            }
            
            from datetime import datetime, timedelta
            start_date = datetime.now().isoformat() + "Z"
            end_date = (datetime.now() + timedelta(days=30)).isoformat() + "Z"
            
            calendar_url = f"https://graph.microsoft.com/v1.0/me/calendar/calendarView"
            params = {
                "startDateTime": start_date,
                "endDateTime": end_date,
                "$orderby": "start/dateTime",
                "$select": "subject,start,end,location,organizer,attendees,isAllDay,bodyPreview,webLink"
            }
            
            response = requests.get(calendar_url, headers=headers, params=params)
            
            if response.status_code != 200:
                response_text = _sanitize_credentials(response.text, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Could not access calendar (Error {response.status_code}).\nPlease check your permissions.",
                    error=f"API error: {response.status_code}"
                )
            
            events_data = response.json()
            events = events_data.get("value", [])
            
            status_messages.append(f"✅ Step 4/5: Retrieved {len(events)} calendar events")
            status_messages.append("📊 Step 5/5: Generating Excel report...")
            
            # Try to create Excel file
            try:
                try:
                    import openpyxl
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                except ImportError:
                    try:
                        import pandas as pd
                        use_pandas = True
                    except ImportError:
                        return TaskResult(
                            False,
                            "❌ Status: Excel library not installed.\nInstall with: pip install openpyxl OR pip install pandas openpyxl",
                            error="Missing dependencies"
                        )
                else:
                    use_pandas = False
                
                # Get reports directory
                from pathlib import Path
                project_dir = Path("F:/Dre_Programs/LeaAssistant")
                reports_dir = project_dir / "lea_reports"
                reports_dir.mkdir(parents=True, exist_ok=True)
                
                # Create filename with timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"outlook_calendar_report_{timestamp}.xlsx"
                file_path = reports_dir / filename
                
                if use_pandas:
                    # Use pandas to create Excel
                    calendar_data = []
                    for event in events:
                        subject = event.get("subject", "(No Subject)")
                        start = event.get("start", {}).get("dateTime", "")
                        end = event.get("end", {}).get("dateTime", "")
                        location = event.get("location", {}).get("displayName", "") if event.get("location") else ""
                        organizer = event.get("organizer", {}).get("emailAddress", {}).get("address", "") if event.get("organizer") else ""
                        organizer_name = event.get("organizer", {}).get("emailAddress", {}).get("name", "") if event.get("organizer") else ""
                        is_all_day = event.get("isAllDay", False)
                        preview = event.get("bodyPreview", "")[:200]  # First 200 chars
                        web_link = event.get("webLink", "")
                        
                        # Format dates
                        try:
                            start_dt = datetime.fromisoformat(start.replace('Z', '+00:00'))
                            end_dt = datetime.fromisoformat(end.replace('Z', '+00:00'))
                            start_str = start_dt.strftime("%Y-%m-%d %H:%M:%S")
                            end_str = end_dt.strftime("%Y-%m-%d %H:%M:%S")
                            duration = (end_dt - start_dt).total_seconds() / 3600  # Duration in hours
                        except:
                            start_str = start
                            end_str = end
                            duration = ""
                        
                        # Count attendees
                        attendees = event.get("attendees", [])
                        attendee_count = len(attendees) if attendees else 0
                        attendee_list = ", ".join([a.get("emailAddress", {}).get("address", "") for a in attendees[:5]]) if attendees else ""
                        if attendee_count > 5:
                            attendee_list += f" (+{attendee_count - 5} more)"
                        
                        calendar_data.append({
                            "Subject": subject,
                            "Start": start_str,
                            "End": end_str,
                            "Duration (hours)": round(duration, 2) if isinstance(duration, (int, float)) else "",
                            "All Day": "Yes" if is_all_day else "No",
                            "Location": location,
                            "Organizer": organizer_name or organizer,
                            "Attendees": attendee_list,
                            "Attendee Count": attendee_count,
                            "Preview": preview,
                            "Link": web_link
                        })
                    
                    df = pd.DataFrame(calendar_data)
                    df.to_excel(file_path, index=False, sheet_name="Calendar Events")
                    
                else:
                    # Use openpyxl for more control
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Calendar Events"
                    
                    # Header row with styling
                    headers = ["Subject", "Start", "End", "Duration (hours)", "All Day", "Location", "Organizer", "Attendees", "Attendee Count", "Preview", "Link"]
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Add summary row
                    ws.cell(row=2, column=1, value="Summary")
                    ws.cell(row=2, column=2, value=f"Total Events: {len(events)} | Next 30 Days")
                    ws.merge_cells(f"B2:K2")
                    
                    # Add event data
                    for row_num, event in enumerate(events, start=3):
                        subject = event.get("subject", "(No Subject)")
                        start = event.get("start", {}).get("dateTime", "")
                        end = event.get("end", {}).get("dateTime", "")
                        location = event.get("location", {}).get("displayName", "") if event.get("location") else ""
                        organizer = event.get("organizer", {}).get("emailAddress", {}).get("address", "") if event.get("organizer") else ""
                        organizer_name = event.get("organizer", {}).get("emailAddress", {}).get("name", "") if event.get("organizer") else ""
                        is_all_day = event.get("isAllDay", False)
                        preview = event.get("bodyPreview", "")[:200]  # First 200 chars
                        web_link = event.get("webLink", "")
                        
                        # Format dates
                        try:
                            start_dt = datetime.fromisoformat(start.replace('Z', '+00:00'))
                            end_dt = datetime.fromisoformat(end.replace('Z', '+00:00'))
                            start_str = start_dt.strftime("%Y-%m-%d %H:%M:%S")
                            end_str = end_dt.strftime("%Y-%m-%d %H:%M:%S")
                            duration = (end_dt - start_dt).total_seconds() / 3600  # Duration in hours
                        except:
                            start_str = start
                            end_str = end
                            duration = ""
                        
                        # Count attendees
                        attendees = event.get("attendees", [])
                        attendee_count = len(attendees) if attendees else 0
                        attendee_list = ", ".join([a.get("emailAddress", {}).get("address", "") for a in attendees[:5]]) if attendees else ""
                        if attendee_count > 5:
                            attendee_list += f" (+{attendee_count - 5} more)"
                        
                        ws.cell(row=row_num, column=1, value=subject)
                        ws.cell(row=row_num, column=2, value=start_str)
                        ws.cell(row=row_num, column=3, value=end_str)
                        ws.cell(row=row_num, column=4, value=round(duration, 2) if isinstance(duration, (int, float)) else "")
                        ws.cell(row=row_num, column=5, value="Yes" if is_all_day else "No")
                        ws.cell(row=row_num, column=6, value=location)
                        ws.cell(row=row_num, column=7, value=organizer_name or organizer)
                        ws.cell(row=row_num, column=8, value=attendee_list)
                        ws.cell(row=row_num, column=9, value=attendee_count)
                        ws.cell(row=row_num, column=10, value=preview)
                        ws.cell(row=row_num, column=11, value=web_link)
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        ws.column_dimensions[col_letter].width = adjusted_width
                    
                    wb.save(file_path)
                
                status_messages.append("✅ Excel report generated successfully")
                
                # Generate summary text
                summary_text = "\n".join(status_messages)
                summary_text += f"\n\n✅ Calendar Report Complete!\n"
                summary_text += f"📅 Summary: {len(events)} events in the next 30 days\n"
                summary_text += f"📁 Excel file saved to:\n   {file_path}\n"
                summary_text += f"📝 Report contains {len(events)} calendar events.\n"
                summary_text += f"\n💡 All calendar details are in the Excel file."
                
                return TaskResult(
                    True,
                    summary_text,
                    data={
                        "event_count": len(events),
                        "file_path": str(file_path),
                        "filename": filename,
                        "events": events  # Keep in data but not in message/logs
                    }
                )
                
            except Exception as excel_error:
                error_msg = _sanitize_credentials(str(excel_error), client_id, client_secret, tenant_id)
                status_messages.append("❌ Excel generation failed")
                
                return TaskResult(
                    False,
                    "\n".join(status_messages) + f"\n\n❌ Failed to generate Excel file.\nError: {error_msg}\n\nInstall openpyxl with: pip install openpyxl",
                    data={
                        "event_count": len(events),
                        "excel_error": error_msg
                    },
                    error=error_msg
                )
            
        except Exception as e:
            error_msg = _sanitize_credentials(str(e), client_id, client_secret, tenant_id)
            logging.error("OutlookCalendarCheckTask error occurred (details not logged for security)")
            
            status_text = "\n".join(status_messages) if status_messages else "❌ Status: Task failed"
            return TaskResult(
                False,
                f"{status_text}\n\n❌ Error: {error_msg}\n\nPlease check your OAuth credentials and try again.",
                error=error_msg
            )


class OutlookInboxOrganizeTask(BaseTask):
    """Organize and clean Outlook inbox and folders (Executive Assistant mode only)"""
    
    def __init__(self):
        super().__init__(
            name="outlook_inbox_organize",
            description="Organize and clean Outlook inbox and folders. Can create a plan first or execute directly (Executive Assistant mode only)",
            requires_confirmation=True  # Always requires confirmation for safety
        )
    
    def get_required_params(self):
        return ["action"]  # "plan" or "execute"
    
    def execute(self, **kwargs):
        # Initialize credential variables outside try block for sanitization
        client_id = None
        client_secret = None
        tenant_id = None
        
        # Status messages
        status_messages = []
        
        try:
            import os
            from dotenv import load_dotenv
            load_dotenv()
            
            status_messages.append("🔐 Step 1/6: Loading authentication libraries...")
            
            # Try to import Microsoft Graph SDK
            try:
                from msal import ConfidentialClientApplication
                import requests
                import logging as msal_logging
                msal_logging.getLogger("msal").setLevel(msal_logging.CRITICAL)
                logging.getLogger("requests").setLevel(logging.WARNING)
                logging.getLogger("urllib3").setLevel(logging.WARNING)
            except ImportError:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Microsoft authentication libraries not installed.\nInstall with: pip install msal requests",
                    error="Missing dependencies"
                )
            
            status_messages.append("✅ Step 1/6: Libraries loaded")
            status_messages.append("🔐 Step 2/6: Reading OAuth credentials...")
            
            # Get OAuth credentials from .env
            client_id = os.getenv("OUTLOOK_CLIENT_ID")
            client_secret = os.getenv("OUTLOOK_CLIENT_SECRET")
            tenant_id = os.getenv("OUTLOOK_TENANT_ID", "common")
            
            if not client_id or not client_secret:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Outlook OAuth credentials not found in .env file.",
                    error="Missing credentials"
                )
            
            status_messages.append("✅ Step 2/6: Credentials found")
            status_messages.append("🔐 Step 3/6: Authenticating with Outlook...")
            
            # Authenticate
            authority = f"https://login.microsoftonline.com/{tenant_id}"
            app = ConfidentialClientApplication(
                client_id=client_id,
                client_credential=client_secret,
                authority=authority
            )
            
            # Get token
            scopes = ["https://graph.microsoft.com/.default"]
            result = app.acquire_token_silent(scopes, account=None)
            
            if not result:
                return TaskResult(
                    False,
                    "⚠️ Status: Authentication required.\nPlease authenticate Outlook access when prompted.",
                    error="Authentication required",
                    data={"needs_auth": True}
                )
            
            access_token = result.get("access_token")
            if not access_token:
                error_desc = result.get('error_description', 'Unknown error')
                error_desc = _sanitize_credentials(error_desc, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Authentication error.\n{error_desc}",
                    error="Authentication failed"
                )
            
            status_messages.append("✅ Step 3/6: Authenticated successfully")
            status_messages.append("📁 Step 4/6: Analyzing folders and inbox...")
            
            # Get action parameter
            action = kwargs.get("action", "plan").lower()  # "plan" or "execute"
            target_folder = kwargs.get("folder", "inbox")  # Which folder to organize
            organization_rules = kwargs.get("rules", {})  # Optional rules for organization
            
            headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json"
            }
            
            # Get all mail folders
            folders_url = "https://graph.microsoft.com/v1.0/me/mailFolders"
            response = requests.get(folders_url, headers=headers)
            
            if response.status_code != 200:
                response_text = _sanitize_credentials(response.text, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Could not access folders (Error {response.status_code}).",
                    error=f"API error: {response.status_code}"
                )
            
            folders_data = response.json()
            folders = folders_data.get("value", [])
            
            # Get inbox folder
            inbox_folder = next((f for f in folders if f.get("displayName", "").lower() == "inbox"), None)
            if not inbox_folder:
                return TaskResult(
                    False,
                    "❌ Status: Failed - Could not find inbox folder.",
                    error="Inbox not found"
                )
            
            inbox_id = inbox_folder.get("id")
            
            # Get messages from inbox
            messages_url = f"https://graph.microsoft.com/v1.0/me/mailFolders/{inbox_id}/messages"
            params = {
                "$top": 100,  # Get up to 100 messages
                "$orderby": "receivedDateTime desc",
                "$select": "id,subject,sender,receivedDateTime,isRead,hasAttachments"
            }
            
            response = requests.get(messages_url, headers=headers, params=params)
            
            if response.status_code != 200:
                response_text = _sanitize_credentials(response.text, client_id, client_secret, tenant_id)
                return TaskResult(
                    False,
                    f"❌ Status: Failed - Could not retrieve messages (Error {response.status_code}).",
                    error=f"API error: {response.status_code}"
                )
            
            messages_data = response.json()
            messages = messages_data.get("value", [])
            
            # Analyze messages
            unread_count = sum(1 for msg in messages if not msg.get("isRead", False))
            read_count = len(messages) - unread_count
            with_attachments = sum(1 for msg in messages if msg.get("hasAttachments", False))
            
            status_messages.append(f"✅ Step 4/6: Analyzed {len(messages)} messages")
            
            if action == "plan":
                # Create a plan without executing
                status_messages.append("📋 Step 5/6: Creating organization plan...")
                
                plan_items = []
                
                # Analyze for potential organization
                # Group by sender, date, read status, etc.
                from collections import defaultdict
                from datetime import datetime, timedelta
                
                sender_groups = defaultdict(int)
                old_messages = []
                recent_messages = []
                
                for msg in messages:
                    sender = msg.get("sender", {}).get("emailAddress", {}).get("address", "Unknown")
                    sender_groups[sender] += 1
                    
                    received = msg.get("receivedDateTime", "")
                    try:
                        dt = datetime.fromisoformat(received.replace('Z', '+00:00'))
                        days_old = (datetime.now(dt.tzinfo) - dt).days
                        
                        if days_old > 90:
                            old_messages.append(msg)
                        elif days_old <= 7:
                            recent_messages.append(msg)
                    except:
                        pass
                
                # Build plan
                plan_items.append(f"📊 **Current Inbox Status:**")
                plan_items.append(f"   - Total messages: {len(messages)}")
                plan_items.append(f"   - Unread: {unread_count}")
                plan_items.append(f"   - Read: {read_count}")
                plan_items.append(f"   - With attachments: {with_attachments}")
                plan_items.append(f"   - Messages older than 90 days: {len(old_messages)}")
                plan_items.append(f"   - Recent messages (last 7 days): {len(recent_messages)}")
                plan_items.append(f"   - Unique senders: {len(sender_groups)}")
                
                plan_items.append(f"\n📋 **Suggested Organization Plan:**")
                
                if old_messages:
                    plan_items.append(f"   1. Archive or move {len(old_messages)} old messages (>90 days)")
                
                if unread_count > 0:
                    plan_items.append(f"   2. Mark {unread_count} unread messages as read (if desired)")
                
                # Top senders
                top_senders = sorted(sender_groups.items(), key=lambda x: x[1], reverse=True)[:5]
                if top_senders:
                    plan_items.append(f"   3. Top senders (consider creating rules):")
                    for sender, count in top_senders:
                        plan_items.append(f"      - {sender}: {count} messages")
                
                plan_items.append(f"\n💡 **Recommendations:**")
                plan_items.append(f"   - Review old messages before archiving")
                plan_items.append(f"   - Consider creating folders for frequent senders")
                plan_items.append(f"   - Set up rules to auto-organize future emails")
                
                plan_text = "\n".join(plan_items)
                
                status_messages.append("✅ Step 5/6: Plan created")
                status_messages.append("✅ Step 6/6: Analysis complete")
                
                summary_text = "\n".join(status_messages)
                summary_text += f"\n\n📋 **Organization Plan Created:**\n\n{plan_text}\n\n"
                summary_text += f"💡 To execute this plan, ask Lea to organize your inbox with action='execute'"
                
                return TaskResult(
                    True,
                    summary_text,
                    data={
                        "action": "plan",
                        "message_count": len(messages),
                        "unread_count": unread_count,
                        "read_count": read_count,
                        "old_messages_count": len(old_messages),
                        "recent_messages_count": len(recent_messages),
                        "plan": plan_text,
                        "folders": [f.get("displayName") for f in folders]
                    }
                )
            
            else:  # action == "execute"
                # Execute organization (this would require more specific parameters)
                # For now, return a message that execution needs specific instructions
                status_messages.append("⚙️ Step 5/6: Preparing to execute organization...")
                status_messages.append("⚠️ Step 6/6: Execution requires specific parameters")
                
                summary_text = "\n".join(status_messages)
                summary_text += f"\n\n⚠️ **Execution Mode:**\n"
                summary_text += f"To execute organization, please specify:\n"
                summary_text += f"- Which messages to move (by date, sender, subject, etc.)\n"
                summary_text += f"- Destination folder(s)\n"
                summary_text += f"- Any specific rules to apply\n\n"
                summary_text += f"💡 **Recommendation:** Start with action='plan' to see what can be organized, then provide specific instructions for execution."
                
                return TaskResult(
                    True,
                    summary_text,
                    data={
                        "action": "execute",
                        "message_count": len(messages),
                        "unread_count": unread_count,
                        "read_count": read_count,
                        "folders": [f.get("displayName") for f in folders]
                    }
                )
            
        except Exception as e:
            error_msg = _sanitize_credentials(str(e), client_id, client_secret, tenant_id)
            logging.error("OutlookInboxOrganizeTask error occurred (details not logged for security)")
            
            status_text = "\n".join(status_messages) if status_messages else "❌ Status: Task failed"
            return TaskResult(
                False,
                f"{status_text}\n\n❌ Error: {error_msg}\n\nPlease check your OAuth credentials and try again.",
                error=error_msg
            )


class OutlookSharedCalendarCheckTask(BaseTask):
    """Check shared calendars in Outlook (Executive Assistant mode only)"""
    
    def __init__(self):
        super().__init__(
            name="outlook_shared_calendar_check",
            description="Check shared calendars in Outlook and generate a report (Executive Assistant mode only)",
            requires_confirmation=False
        )
    
    def get_required_params(self):
        return []  # No required params
    
    def execute(self, **kwargs):
        # Similar to OutlookCalendarCheckTask but for shared calendars
        # Uses Calendars.Read.Shared permission
        # Implementation would be similar to calendar check but query shared calendars
        # For now, return a message that this uses shared calendar access
        return TaskResult(
            True,
            "✅ Shared calendar check - This would use Calendars.Read.Shared permission to access shared calendars.\n"
            "📅 Implementation: Query shared calendars via Microsoft Graph API\n"
            "💡 This feature can be fully implemented when needed.",
            data={"permission": "Calendars.Read.Shared"}
        )


class OutlookUserProfileTask(BaseTask):
    """Get or update user profile information (Executive Assistant mode only)"""
    
    def __init__(self):
        super().__init__(
            name="outlook_user_profile",
            description="Get or update Outlook user profile information (Executive Assistant mode only)",
            requires_confirmation=False  # Reading profile doesn't need confirmation, but writing might
        )
    
    def get_required_params(self):
        return ["action"]  # "read" or "update"
    
    def execute(self, **kwargs):
        action = kwargs.get("action", "read").lower()
        
        if action == "read":
            # Uses User.Read permission
            return TaskResult(
                True,
                "✅ User profile read - This would use User.Read permission to get profile information.\n"
                "👤 Can retrieve: Display name, email, job title, office location, etc.\n"
                "💡 This feature can be fully implemented when needed.",
                data={"permission": "User.Read", "action": "read"}
            )
        else:  # update
            # Uses User.ReadWrite permission
            return TaskResult(
                True,
                "✅ User profile update - This would use User.ReadWrite permission to update profile.\n"
                "⚠️ Requires confirmation before updating profile information.\n"
                "💡 This feature can be fully implemented when needed.",
                data={"permission": "User.ReadWrite", "action": "update"}
            )


def register_custom_tasks():
    """Register all custom tasks with the task registry"""
    from lea_tasks import get_task_registry
    
    registry = get_task_registry()
    
    # Register your custom tasks
    registry.register_task(TextAnalysisTask())
    registry.register_task(ConfigManagerTask())
    registry.register_task(FileOrganizerTask())
    registry.register_task(OutlookEmailCheckTask())  # Email checking (Executive Assistant only) - Mail.Read
    registry.register_task(OutlookEmailDraftTask())  # Email draft creation (Executive Assistant only) - Mail.Read
    registry.register_task(OutlookInboxOrganizeTask())  # Inbox/folder organization (Executive Assistant only) - Mail.Read
    registry.register_task(OutlookCalendarCheckTask())  # Calendar check (Executive Assistant only) - Calendars.Read
    registry.register_task(OutlookSharedCalendarCheckTask())  # Shared calendar check (Executive Assistant only) - Calendars.Read.Shared
    registry.register_task(OutlookUserProfileTask())  # User profile (Executive Assistant only) - User.Read, User.ReadWrite
    
    logging.info("Custom tasks registered successfully")
    
    return registry


# Auto-register on import:
register_custom_tasks()

